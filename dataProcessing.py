from __future__ import print_function
import time
import swagger_client
import requests
import json
import openpyxl
import math
import schedule
import time
from bullet import VerticalPrompt, Password, Input
from pathlib import Path
from swagger_client.rest import ApiException
from pprint import pprint
from requests.auth import HTTPBasicAuth
from influxdb import InfluxDBClient
from datetime import datetime

# -------------------------------- Saving data from the xlsx file to a Dict ----------------------------------

# Function to read the data given from the xlsx file
def readXlsx(dir, fileName):
    # Setting the path to the xlsx file:
    xlsx_file = Path(dir, fileName)

    # Read the Excel File
    wb_obj = openpyxl.load_workbook(xlsx_file)

    # Read the Active Sheet from the Excel file
    sheet = wb_obj.active

    # Max rows
    # print(("Number of rows: %d") % (sheet.max_row))

    accessPoints = {}
    i = 0
    for row in sheet.iter_rows(max_col=7, values_only=True):
        if (i != 0):
            id = row[0]
            apData = {
                'location' : row[1],
                'name' : row[2],
                'latitude' : row[3],
                'longitude' : row[4],
                'responsible' : row[5],
                'building' : row[6] 
            }
            accessPoints[id] = apData
        i+=1

    return accessPoints

xlsxData = readXlsx('.', 'PrimeCore.xlsx')

# ------------------------------------------- Functions ------------------------------------------------------

# Function to get the API access token (expires each hour)
def getAPIAccessToken():
    global api_instance 
    global tokenExpiresIn 
    print("Calling token")

    # Getting the access token
    url = 'https://wso2-gw.ua.pt/token?grant_type=client_credentials&state=123&scope=openid'
    header = {'Content-Type': 'application/x-www-form-urlencoded'}

    x = requests.post(url,headers=header,auth=HTTPBasicAuth('***REMOVED***','***REMOVED***'))
    resp = x.json()

    # Configure OAuth2 access token for authorization
    swagger_client.configuration.access_token = resp["access_token"]

    tokenExpiresIn = resp["expires_in"]

    print("token: "+str(swagger_client.configuration.access_token))

    # create an instance of the API class
    api_instance = swagger_client.DefaultApi()  

    # print("expires in: "+str(tokenExpiresIn))

# Function to create the database
def createDB():
    cli = VerticalPrompt([
        Input(prompt="username: "),
        Password(prompt="password: ", hidden = "*")],
        spacing=0)
    
    result = cli.launch()

    client = InfluxDBClient("localhost", 8086, str(result[0][1]), str(result[1][1]), "***REMOVED***")
    client.create_database("***REMOVED***")
    client.get_list_database()
    client.switch_database("***REMOVED***")

    return client

# Function to get access points Info
def getAccessPoints(client, numReq):
    apInfo = []
    numReq = numReq * 100;

    api_response = api_instance.access_point_get(first_result=numReq)

    # Get the first index
    firstIndex = int(api_response.first)

    # Get the last index
    lastIndex = int(api_response.last)

    # Get the access points list
    resp = api_response.access_points

    stackSize = lastIndex - firstIndex
    for i in range(0, stackSize + 1):
        apInfo.append(int(resp[i].id))
        apInfo.append(resp[i].name)    

        # Get the building by the id
        building = xlsxData[apInfo[0]].get('building')
        apInfo.append(building)

        apInfo.append(int(resp[i].client_count))    
        apInfo.append(int(resp[i].client_count_2_4_g_hz))
        apInfo.append(int(resp[i].client_count_5_g_hz))    

        # Write on database
        writeAccessPointsOnDB(client, apInfo)

        # Clearing the list of access points Info
        apInfo.clear()

# Function to write the access points Info on the database
def writeAccessPointsOnDB(client, info):
    # Data to send to the database
    json_payload = []

    #
    # Note: tags -> metadata about the measurement 
    #       fields -> a measurement that changes over time
    #

    data = { 
        "measurement" : "clientsCount",
        "time" : datetime.now(),
        "tags" : {
            "id" : info[0],
            "name" : info[1],
            "building" : info[2]
        },
        "fields" : {
            "clientsCount" : info[3],
            "clientsCount2_4Ghz" : info[4],
            "clientsCount5Ghz" : info[5]
        }
    }

    # Send data to the API
    json_payload.append(data)
    client.write_points(json_payload)

# Function to call the API to get the access points
def apiGetAccessPoint(client):
    print("Calling Access Points")

    # To get the total number of working access points
    apsCount = int(api_instance.access_point_count_get().count)

    # Defining the number os API requests needed
    numberReq = math.ceil(apsCount / 100)

    for index in range (0, numberReq):
        # Calling function to get the access points info
        getAccessPoints(client, index)

# ------------------------------------------ Main Function --------------------------------------------------- 

# Creating the database
client = createDB()

# Getting the first data from the API
getAPIAccessToken()

min = math.ceil(tokenExpiresIn / 60)
schedule.every(min).minutes.do(getAPIAccessToken)

apiGetAccessPoint(client)

# Calling the API to get the access token every hour
#schedule.every(5).seconds.do(getAPIAccessToken)

# Calling the API to get the access points every 15 minutes
schedule.every(15).minutes.do(apiGetAccessPoint, client)

while True:
    try:
        schedule.run_pending()
    except ApiException as e:
        print("Exception: %s\n" % e)